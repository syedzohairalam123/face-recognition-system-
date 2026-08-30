"""
Attendance Routes
-----------------
Routes for attendance management, check-in/out, and reports.
Includes real-time recognition endpoint for Phase 10.
"""

import logging
import os
import uuid
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session

from app.services.attendance_service import AttendanceService
from app.services.user_service import UserService
from app.utils.helpers import get_status_color, format_datetime, format_duration
from app.utils.decorators import login_required, admin_required
from config.settings import UPLOAD_DIR

logger = logging.getLogger(__name__)
attendance_bp = Blueprint("attendance", __name__)


@attendance_bp.route("/")
def today():
    """View today's attendance records."""
    service = AttendanceService()
    records = service.get_today_records()
    stats = service.get_attendance_stats()

    records_data = []
    for record in records:
        data = record.to_dict()
        data["status_color"] = get_status_color(record.status)
        records_data.append(data)

    users = UserService.get_all_users(active_only=True)

    return render_template(
        "attendance/today.html",
        records=records_data,
        stats=stats,
        users=users,
        format_datetime=format_datetime,
        format_duration=format_duration,
    )


@attendance_bp.route("/history")
def history():
    """View attendance history with advanced filtering, search, and sorting."""
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to", datetime.now().strftime("%Y-%m-%d"))
    user_filter = request.args.get("user_id", type=int)
    status_filter = request.args.get("status", "all")
    department_filter = request.args.get("department", "all")
    search_query = request.args.get("search", "").strip()
    sort_by = request.args.get("sort", "date_desc")

    service = AttendanceService()

    if date_from:
        records = service.get_records_by_date_range(date_from, date_to)
    else:
        records = service.get_records_by_date(date_to)

    # Apply filters
    if user_filter:
        records = [r for r in records if r.user_id == user_filter]

    if status_filter and status_filter != "all":
        records = [r for r in records if r.status == status_filter]

    if department_filter and department_filter != "all":
        records = [r for r in records if r.user and r.user.department == department_filter]

    if search_query:
        search_lower = search_query.lower()
        records = [
            r for r in records
            if r.user and (
                search_lower in (r.user.employee_id or "").lower()
                or search_lower in (r.user.full_name or "").lower()
                or search_lower in (r.user.department or "").lower()
            )
        ]

    # Sort records
    reverse = True  # default descending
    if sort_by == "date_asc":
        records.sort(key=lambda r: r.check_in_time or datetime.min, reverse=False)
        reverse = False
    elif sort_by == "date_desc":
        records.sort(key=lambda r: r.check_in_time or datetime.min, reverse=True)
    elif sort_by == "name_asc":
        records.sort(key=lambda r: r.user.full_name if r.user else "", reverse=False)
        reverse = False
    elif sort_by == "name_desc":
        records.sort(key=lambda r: r.user.full_name if r.user else "", reverse=True)
    elif sort_by == "status_asc":
        records.sort(key=lambda r: r.status or "", reverse=False)
        reverse = False
    elif sort_by == "status_desc":
        records.sort(key=lambda r: r.status or "", reverse=True)

    records_data = []
    for record in records:
        data = record.to_dict()
        data["status_color"] = get_status_color(record.status)
        records_data.append(data)

    # Get filter options
    all_users = UserService.get_all_users(active_only=True)
    departments = list(set(u.department for u in all_users if u.department))
    departments.sort()

    result_count = len(records_data)

    return render_template(
        "attendance/history.html",
        records=records_data,
        selected_date=date_to,
        date_from=date_from,
        date_to=date_to,
        user_filter=user_filter,
        status_filter=status_filter,
        department_filter=department_filter,
        search_query=search_query,
        sort_by=sort_by,
        all_users=all_users,
        departments=departments,
        result_count=result_count,
        format_datetime=format_datetime,
        format_duration=format_duration,
    )


@attendance_bp.route("/check-in", methods=["POST"])
@admin_required
def check_in():
    """Manual check-in for a user (admin only)."""
    user_id = request.form.get("user_id", type=int)

    if not user_id:
        flash("Please select a user.", "warning")
        return redirect(url_for("attendance.today"))

    service = AttendanceService()
    success, message = service.mark_attendance(user_id)

    flash(message, "success" if success else "warning")
    return redirect(url_for("attendance.today"))


@attendance_bp.route("/check-out", methods=["POST"])
@admin_required
def check_out():
    """Manual check-out for a user (admin only)."""
    user_id = request.form.get("user_id", type=int)

    if not user_id:
        flash("Please select a user.", "warning")
        return redirect(url_for("attendance.today"))

    service = AttendanceService()
    success, message = service.check_out(user_id)

    flash(message, "success" if success else "warning")
    return redirect(url_for("attendance.today"))


@attendance_bp.route("/recognize", methods=["POST"])
@login_required
def recognize_and_mark():
    """
    Recognize a face from uploaded image and mark attendance.
    Used by the camera/upload interface.
    """
    file = request.files.get("face_image")
    if not file or file.filename == "":
        return jsonify({"success": False, "message": "No image provided"}), 400

    # Save uploaded image
    ext = file.filename.rsplit(".", 1)[1].lower() if "." in file.filename else "jpg"
    unique_name = f"capture_{uuid.uuid4().hex[:8]}.{ext}"
    filepath = os.path.join(str(UPLOAD_DIR), unique_name)
    file.save(filepath)

    try:
        from app.services.recognition_pipeline import RecognitionPipeline

        pipeline = RecognitionPipeline(auto_mark_attendance=True)
        results = pipeline.process_image_file(filepath, mark_attendance=True)

        if not results:
            return jsonify({
                "success": False,
                "message": "No faces detected in image",
            }), 404

        result = results[0]

        if not result.success:
            return jsonify({
                "success": False,
                "message": "Face not recognized",
                "metrics": result.metrics,
            }), 404

        # Include decision engine results if available
        decision_result = None
        if hasattr(result, 'decision_result') and result.decision_result:
            decision_result = result.decision_result.to_dict()

        return jsonify({
            "success": True,
            "message": result.attendance_message or f"Recognized {result.full_name}",
            "employee_id": result.employee_id,
            "full_name": result.full_name,
            "confidence": result.confidence,
            "user_id": result.user_id,
            "attendance_marked": result.attendance_marked,
            "processing_time_ms": result.processing_time_ms,
            "metrics": result.metrics,
            "decision_result": decision_result,
        })

    except ImportError as e:
        return jsonify({
            "success": False,
            "message": f"Required library not installed: {e}",
        }), 503
    except Exception as e:
        logger.error(f"Recognition error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@attendance_bp.route("/recognize-realtime", methods=["POST"])
@login_required
def recognize_realtime():
    """
    Real-time recognition endpoint for live camera feed.
    Accepts base64 encoded image from video stream.
    
    Phase 10: Live Face Recognition UI
    """
    data = request.get_json()
    if not data or "image" not in data:
        return jsonify({"success": False, "message": "No image data provided"}), 400

    import base64
    import cv2
    import numpy as np

    try:
        # Decode base64 image
        image_data = data["image"].split(",")[1] if "," in data["image"] else data["image"]
        image_bytes = base64.b64decode(image_data)
        nparr = np.frombuffer(image_bytes, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

        if frame is None:
            return jsonify({"success": False, "message": "Could not decode image"}), 400

        from app.services.recognition_pipeline import RecognitionPipeline

        pipeline = RecognitionPipeline(auto_mark_attendance=False)
        results = pipeline.process_frame(frame, mark_attendance=False)

        # Format results for UI
        faces = []
        for result in results:
            face_data = {
                "user_id": result.user_id,
                "employee_id": result.employee_id,
                "full_name": result.full_name,
                "confidence": result.confidence,
                "success": result.success,
            }
            if result.metrics and "face_location" in result.metrics:
                face_data["face_location"] = result.metrics["face_location"]
            
            # Include decision engine results if available
            if hasattr(result, 'decision_result') and result.decision_result:
                face_data["decision_result"] = result.decision_result.to_dict()
            
            faces.append(face_data)

        return jsonify({
            "success": True,
            "faces": faces,
            "face_count": len(faces),
            "processing_time_ms": results[0].processing_time_ms if results else 0,
        })

    except Exception as e:
        logger.error(f"Real-time recognition error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@attendance_bp.route("/stats")
def stats():
    """Get attendance statistics as JSON."""
    date_str = request.args.get("date")
    service = AttendanceService()
    stats = service.get_attendance_stats(date_str)
    return jsonify(stats)


@attendance_bp.route("/daily-report")
def daily_report():
    """Get daily report as JSON."""
    days = request.args.get("days", 7, type=int)
    service = AttendanceService()
    report = service.get_daily_report(days)
    return jsonify(report)


@attendance_bp.route("/export-csv")
def export_csv():
    """Export attendance records as CSV file (respects filters)."""
    import csv
    import io
    from flask import Response

    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to", datetime.now().strftime("%Y-%m-%d"))
    user_filter = request.args.get("user_id", type=int)
    status_filter = request.args.get("status", "all")
    department_filter = request.args.get("department", "all")
    search_query = request.args.get("search", "").strip()

    service = AttendanceService()

    if date_from:
        records = service.get_records_by_date_range(date_from, date_to)
    else:
        records = service.get_records_by_date(date_to)

    # Apply same filters as history page
    if user_filter:
        records = [r for r in records if r.user_id == user_filter]
    if status_filter and status_filter != "all":
        records = [r for r in records if r.status == status_filter]
    if department_filter and department_filter != "all":
        records = [r for r in records if r.user and r.user.department == department_filter]
    if search_query:
        search_lower = search_query.lower()
        records = [
            r for r in records
            if r.user and (
                search_lower in (r.user.employee_id or "").lower()
                or search_lower in (r.user.full_name or "").lower()
                or search_lower in (r.user.department or "").lower()
            )
        ]

    # Build filename
    parts = ["attendance"]
    if date_from:
        parts.append(f"{date_from}_to_{date_to}")
    else:
        parts.append(date_to)
    if status_filter and status_filter != "all":
        parts.append(status_filter)
    if department_filter and department_filter != "all":
        parts.append(department_filter.replace(" ", "_"))
    filename = "_".join(parts) + ".csv"

    # Create CSV content
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        "Employee ID", "Employee Name", "Department", "Date",
        "Check In", "Check Out", "Duration (min)", "Status", "Confidence"
    ])

    # Data rows
    for record in records:
        user = record.user
        duration = record.duration if record.duration else 0
        confidence = f"{record.confidence_score * 100:.1f}%" if record.confidence_score else "-"

        writer.writerow([
            user.employee_id if user else "-",
            user.full_name if user else "-",
            user.department if user and user.department else "-",
            record.attendance_date.strftime("%Y-%m-%d") if record.attendance_date else "-",
            record.check_in_time.strftime("%H:%M:%S") if record.check_in_time else "-",
            record.check_out_time.strftime("%H:%M:%S") if record.check_out_time else "-",
            f"{duration:.0f}" if duration else "-",
            record.status.replace("_", " ").title() if record.status else "-",
            confidence,
        ])

    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@attendance_bp.route("/export-excel")
def export_excel():
    """Export attendance records as Excel (.xlsx) file (respects filters)."""
    from flask import Response
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("Excel export requires openpyxl. Install it with: pip install openpyxl", "warning")
        return redirect(url_for("attendance.history"))

    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to", datetime.now().strftime("%Y-%m-%d"))
    user_filter = request.args.get("user_id", type=int)
    status_filter = request.args.get("status", "all")
    department_filter = request.args.get("department", "all")
    search_query = request.args.get("search", "").strip()

    service = AttendanceService()

    if date_from:
        records = service.get_records_by_date_range(date_from, date_to)
    else:
        records = service.get_records_by_date(date_to)

    # Apply filters
    if user_filter:
        records = [r for r in records if r.user_id == user_filter]
    if status_filter and status_filter != "all":
        records = [r for r in records if r.status == status_filter]
    if department_filter and department_filter != "all":
        records = [r for r in records if r.user and r.user.department == department_filter]
    if search_query:
        search_lower = search_query.lower()
        records = [
            r for r in records
            if r.user and (
                search_lower in (r.user.employee_id or "").lower()
                or search_lower in (r.user.full_name or "").lower()
                or search_lower in (r.user.department or "").lower()
            )
        ]

    # Build filename
    parts = ["attendance"]
    if date_from:
        parts.append(f"{date_from}_to_{date_to}")
    else:
        parts.append(date_to)
    if status_filter and status_filter != "all":
        parts.append(status_filter)
    filename = "_".join(parts) + ".xlsx"

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    status_fills = {
        "present": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        "late": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "checked_out": PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid"),
        "absent": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
    }

    # Title row
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"Attendance Report — {date_from or 'Single Day'} to {date_to}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="2B2D42")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Summary row
    ws.merge_cells("A2:I2")
    summary_cell = ws["A2"]
    summary_cell.value = f"Total Records: {len(records)} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    summary_cell.font = Font(name="Calibri", italic=True, size=10, color="6C757D")
    ws.row_dimensions[2].height = 20

    # Headers
    headers = [
        "Employee ID", "Employee Name", "Department", "Date",
        "Check In", "Check Out", "Duration (min)", "Status", "Confidence"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[4].height = 25

    # Data rows
    for row_num, record in enumerate(records, 5):
        user = record.user
        duration = record.duration if record.duration else 0
        confidence = f"{record.confidence_score * 100:.1f}%" if record.confidence_score else "-"

        row_data = [
            user.employee_id if user else "-",
            user.full_name if user else "-",
            user.department if user and user.department else "-",
            record.attendance_date.strftime("%Y-%m-%d") if record.attendance_date else "-",
            record.check_in_time.strftime("%H:%M:%S") if record.check_in_time else "-",
            record.check_out_time.strftime("%H:%M:%S") if record.check_out_time else "-",
            f"{duration:.0f}" if duration else "-",
            record.status.replace("_", " ").title() if record.status else "-",
            confidence,
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Color status column
            if col_num == 8:
                status_key = record.status if record else ""
                if status_key in status_fills:
                    cell.fill = status_fills[status_key]

    # Auto-fit column widths
    for col in range(1, len(headers) + 1):
        max_length = len(headers[col - 1])
        for row in range(5, 5 + len(records)):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val:
                max_length = max(max_length, len(str(cell_val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 3, 30)

    # Save to bytes
    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
